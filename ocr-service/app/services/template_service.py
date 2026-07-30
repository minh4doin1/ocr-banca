"""
Template Service — CRUD for TemplateProfile + infer headers from sample files.
"""

from __future__ import annotations

import json
import logging
import re
import uuid
from datetime import datetime
from pathlib import Path

from app.config import settings
from app.models.schemas import (
    TemplateColumn,
    TemplateExportConfig,
    TemplateOcrConfig,
    TemplateProfile,
    TemplateTableConfig,
    TemplateUpdateRequest,
)
from app.services.action_registry import SSO_ACTIONS, normalize_actions

logger = logging.getLogger(__name__)

DEFAULT_TEMPLATE_ID = "sso-agribank"
_ID_RE = re.compile(r"^[a-zA-Z0-9][a-zA-Z0-9_-]{0,63}$")

_KNOWN_INTERNAL_FIELDS = (
    "stt",
    "username",
    "email",
    "name",
    "first_name",
    "last_name",
    "cccd",
    "branch_name",
    "department_name",
    "branch_code",
    "agent_code",
    "ipcas_code",
    "phone",
    "unit_code",
    "role",
    "password",
    "notes",
)


def _profiles_dir() -> Path:
    path = settings.templates_path / "profiles"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _samples_dir() -> Path:
    path = settings.templates_path / "samples"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _profile_path(template_id: str) -> Path:
    return _profiles_dir() / f"{template_id}.json"


def build_sso_builtin_profile() -> TemplateProfile:
    """Built-in SSO Agribank Mẫu 01 (10 cột) — tương thích ngược."""
    columns = [
        TemplateColumn(index=0, header="STT", field="stt", required=False),
        TemplateColumn(index=1, header="Họ và tên", field="name", required=True),
        TemplateColumn(index=2, header="Mã chi nhánh", field="branch_code", required=True),
        TemplateColumn(index=3, header="Tên Chi nhánh", field="branch_name", required=False),
        TemplateColumn(index=4, header="User IPCAS", field="ipcas_code", required=True),
        TemplateColumn(index=5, header="Số CCCD", field="cccd", required=True),
        TemplateColumn(
            index=6, header="Email tại Agribank", field="email", required=True
        ),
        TemplateColumn(index=7, header="SĐT", field="phone", required=True),
        TemplateColumn(index=8, header="Phân quyền", field="role", required=True),
        TemplateColumn(
            index=9, header="Mã liên ngân hàng", field="unit_code", required=False
        ),
    ]
    excel_headers = [
        "STT",
        "Ho va ten",
        "Ma chi nhanh",
        "Ten Chi nhanh",
        "User IPCAS",
        "So CCCD",
        "Email tai Agribank",
        "SDT",
        "Phan quyen",
        "Ma lien ngan hang",
        "Vai tro (goi y)",
    ]
    aliases = {
        field: list(aliases)
        for field, aliases in settings.keycloak_header_map_parsed.items()
    }
    return TemplateProfile(
        id=DEFAULT_TEMPLATE_ID,
        name="SSO Agribank Mẫu 01",
        version=1,
        builtin=True,
        source_sample="",
        table=TemplateTableConfig(
            header_row=0,
            columns=columns,
            header_aliases=aliases,
        ),
        ocr=TemplateOcrConfig(
            expect_min_cols=9,
            sso_enhance=True,
            table_kind="sso_agribank",
        ),
        actions=list(SSO_ACTIONS),
        export=TemplateExportConfig(
            excel_headers=excel_headers,
            docx_style="table",
            docx_title="Danh sách user SSO — Agribank Banca",
        ),
    )


def ensure_builtin_templates() -> None:
    """Persist built-in SSO profile if missing (idempotent)."""
    path = _profile_path(DEFAULT_TEMPLATE_ID)
    if path.exists():
        return
    profile = build_sso_builtin_profile()
    _write_profile(profile)
    logger.info("Created built-in template profile: %s", DEFAULT_TEMPLATE_ID)


def _write_profile(profile: TemplateProfile) -> None:
    path = _profile_path(profile.id)
    payload = profile.model_dump(mode="json")
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _read_profile(path: Path) -> TemplateProfile:
    data = json.loads(path.read_text(encoding="utf-8"))
    return TemplateProfile.model_validate(data)


def list_templates() -> list[TemplateProfile]:
    ensure_builtin_templates()
    profiles: list[TemplateProfile] = []
    for path in sorted(_profiles_dir().glob("*.json")):
        try:
            profiles.append(_read_profile(path))
        except Exception as exc:
            logger.warning("Skip invalid template %s: %s", path.name, exc)
    # Ensure builtin always present even if file corrupt
    ids = {p.id for p in profiles}
    if DEFAULT_TEMPLATE_ID not in ids:
        profiles.insert(0, build_sso_builtin_profile())
    return profiles


def get_template(template_id: str) -> TemplateProfile | None:
    ensure_builtin_templates()
    tid = (template_id or "").strip() or settings.default_template_id
    path = _profile_path(tid)
    if path.exists():
        try:
            return _read_profile(path)
        except Exception as exc:
            logger.warning("Failed to load template %s: %s", tid, exc)
    if tid == DEFAULT_TEMPLATE_ID:
        return build_sso_builtin_profile()
    return None


def get_template_or_default(template_id: str | None = None) -> TemplateProfile:
    profile = get_template(template_id or settings.default_template_id)
    if profile is None:
        return build_sso_builtin_profile()
    return profile


def validate_template_id(template_id: str) -> str:
    tid = (template_id or "").strip() or settings.default_template_id
    if get_template(tid) is None:
        raise ValueError(f"Template không tồn tại: {tid}")
    return tid


def _slugify_id(raw: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9_-]+", "-", (raw or "").strip().lower())
    s = s.strip("-_")[:64]
    if not s or not _ID_RE.match(s):
        s = f"tpl-{uuid.uuid4().hex[:8]}"
    return s


def save_template(profile: TemplateProfile, *, overwrite: bool = False) -> TemplateProfile:
    ensure_builtin_templates()
    tid = profile.id.strip()
    if not _ID_RE.match(tid):
        raise ValueError(
            "template id chỉ gồm chữ, số, gạch ngang/dưới; bắt đầu bằng chữ/số"
        )
    path = _profile_path(tid)
    existing = get_template(tid) if path.exists() else None
    if existing and existing.builtin and not overwrite:
        # Allow update of non-identity fields on builtin via explicit overwrite path
        pass
    if path.exists() and not overwrite and existing and not existing.builtin:
        raise ValueError(f"Template đã tồn tại: {tid}")

    profile.id = tid
    profile.actions = normalize_actions(profile.actions)
    profile.updated_at = datetime.now()
    if existing:
        profile.created_at = existing.created_at
        profile.builtin = existing.builtin
    _write_profile(profile)
    return profile


def update_template(
    template_id: str, request: TemplateUpdateRequest
) -> TemplateProfile:
    profile = get_template(template_id)
    if profile is None:
        raise ValueError(f"Template không tồn tại: {template_id}")

    data = profile.model_dump()
    if request.name is not None:
        data["name"] = request.name
    if request.table is not None:
        data["table"] = request.table.model_dump()
    if request.ocr is not None:
        data["ocr"] = request.ocr.model_dump()
    if request.actions is not None:
        data["actions"] = normalize_actions(request.actions)
    if request.export is not None:
        data["export"] = request.export.model_dump()
    data["updated_at"] = datetime.now()
    updated = TemplateProfile.model_validate(data)
    _write_profile(updated)
    return updated


def delete_template(template_id: str) -> None:
    profile = get_template(template_id)
    if profile is None:
        raise ValueError(f"Template không tồn tại: {template_id}")
    if profile.builtin:
        raise ValueError("Không thể xóa template built-in")
    path = _profile_path(profile.id)
    if path.exists():
        path.unlink()


def list_available_fields() -> list[dict[str, str]]:
    labels = settings.field_labels_vi
    out: list[dict[str, str]] = []
    for field in _KNOWN_INTERNAL_FIELDS:
        out.append(
            {
                "id": field,
                "label": labels.get(field, field),
            }
        )
    return out


def _guess_field_from_header(header: str) -> str:
    """Map header text → internal field using global + common aliases."""
    from app.services.user_mapping import (
        _header_cell_matches_field,
        _normalize_header_key,
    )

    norm = _normalize_header_key(header)
    if not norm:
        return ""
    # STT common
    if norm in ("stt", "st", "so thu tu", "tt"):
        return "stt"
    alias_map = settings.keycloak_header_map_parsed
    for field, aliases in alias_map.items():
        if _header_cell_matches_field(norm, aliases):
            return field
    return ""


def _infer_columns_from_headers(headers: list[str]) -> list[TemplateColumn]:
    columns: list[TemplateColumn] = []
    for idx, header in enumerate(headers):
        text = str(header or "").strip()
        field = _guess_field_from_header(text)
        required = field in settings.user_required_fields_list
        columns.append(
            TemplateColumn(
                index=idx,
                header=text or f"Cột {idx + 1}",
                field=field,
                required=required,
            )
        )
    return columns


def _read_excel_header_row(path: Path) -> tuple[list[str], list[str]]:
    from openpyxl import load_workbook

    warnings: list[str] = []
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            title = (ws.title or "").strip().lower()
            if title == "chú thích":
                continue
            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                # trim trailing empties
                while cells and not cells[-1]:
                    cells.pop()
                non_empty = [c for c in cells if c]
                if len(non_empty) < 2:
                    continue
                # Prefer row that looks like headers (many guessed fields)
                guessed = sum(1 for c in cells if _guess_field_from_header(c))
                if guessed >= 2 or len(non_empty) >= 3:
                    return cells, warnings
        warnings.append("Không tìm thấy dòng header rõ ràng trong Excel")
        return [], warnings
    finally:
        wb.close()


def _read_docx_header_row(path: Path) -> tuple[list[str], list[str]]:
    warnings: list[str] = []
    try:
        from docx import Document
    except ImportError:
        return [], ["Thiếu python-docx để đọc file Word mẫu"]

    doc = Document(str(path))
    for table in doc.tables:
        if not table.rows:
            continue
        cells = [c.text.strip() for c in table.rows[0].cells]
        while cells and not cells[-1]:
            cells.pop()
        if len([c for c in cells if c]) >= 2:
            return cells, warnings
    warnings.append("Không tìm thấy bảng có header trong Word")
    return [], warnings


def _read_pdf_header_row(path: Path) -> tuple[list[str], list[str]]:
    warnings: list[str] = []
    try:
        import pdfplumber
    except ImportError:
        return [], ["Thiếu pdfplumber để đọc PDF mẫu (chỉ PDF có lớp text)"]

    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages[:3]:
                tables = page.extract_tables() or []
                for table in tables:
                    if not table:
                        continue
                    row0 = [str(c or "").strip() for c in table[0]]
                    while row0 and not row0[-1]:
                        row0.pop()
                    if len([c for c in row0 if c]) >= 2:
                        return row0, warnings
    except Exception as exc:
        warnings.append(f"Không đọc được bảng từ PDF: {exc}")
        return [], warnings
    warnings.append(
        "PDF không có lớp text/bảng — nên dùng Excel hoặc Word làm file mẫu"
    )
    return [], warnings


def infer_from_sample(
    sample_path: Path,
    *,
    name: str = "",
    template_id: str = "",
) -> tuple[TemplateProfile, list[str]]:
    """Parse sample file and build a draft TemplateProfile."""
    sample_path = Path(sample_path)
    ext = sample_path.suffix.lower()
    warnings: list[str] = []

    if ext in (".xlsx", ".xlsm"):
        headers, w = _read_excel_header_row(sample_path)
        warnings.extend(w)
    elif ext == ".docx":
        headers, w = _read_docx_header_row(sample_path)
        warnings.extend(w)
    elif ext == ".pdf":
        headers, w = _read_pdf_header_row(sample_path)
        warnings.extend(w)
    else:
        raise ValueError(f"Định dạng mẫu không hỗ trợ: {ext}")

    if not headers:
        raise ValueError(
            "Không suy ra được header từ file mẫu. "
            + ("; ".join(warnings) if warnings else "")
        )

    columns = _infer_columns_from_headers(headers)
    unmapped = [c.header for c in columns if not c.field]
    if unmapped:
        warnings.append(
            "Chưa map được cột: " + ", ".join(unmapped[:8])
            + ("…" if len(unmapped) > 8 else "")
        )

    tid = template_id.strip() or _slugify_id(name or sample_path.stem)
    display_name = name.strip() or sample_path.stem or tid
    excel_headers = [c.header for c in columns]

    profile = TemplateProfile(
        id=tid,
        name=display_name,
        version=1,
        builtin=False,
        source_sample="",
        table=TemplateTableConfig(header_row=0, columns=columns),
        ocr=TemplateOcrConfig(
            expect_min_cols=max(1, len(columns)),
            sso_enhance=False,
            table_kind="",
        ),
        actions=list(normalize_actions(None)),
        export=TemplateExportConfig(
            excel_headers=excel_headers,
            docx_style="table",
            docx_title=display_name,
        ),
    )
    return profile, warnings


def store_sample_file(
    content: bytes,
    filename: str,
    template_id: str,
) -> str:
    """Save sample under storage/templates/samples; return relative path."""
    safe_name = Path(filename).name
    dest = _samples_dir() / f"{template_id}_{safe_name}"
    dest.write_bytes(content)
    return f"samples/{dest.name}"


def col_index_to_field_map(profile: TemplateProfile) -> dict[int, str]:
    """{col_index: field} for columns that have a field mapped."""
    out: dict[int, str] = {}
    for col in profile.table.columns:
        if col.field:
            out[col.index] = col.field
    return out


def field_to_col_index_map(profile: TemplateProfile) -> dict[str, int]:
    """{field: col_index} — first wins if duplicates."""
    out: dict[str, int] = {}
    for col in profile.table.columns:
        if col.field and col.field not in out:
            out[col.field] = col.index
    return out


def export_headers_for(profile: TemplateProfile) -> list[str]:
    if profile.export.excel_headers:
        return list(profile.export.excel_headers)
    if profile.table.columns:
        return [c.header or f"Cột {c.index + 1}" for c in profile.table.columns]
    return []
