"""
Excel Service — Export/import OCR results as Excel files.
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import settings
from app.models.schemas import CellData, OcrResult, PageResult, TableData
from app.services.user_mapping import _normalize_header_key, _strip_sso_preamble_rows

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────
# Styles
# ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(
    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")

LOW_CONFIDENCE_FILL = PatternFill(
    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
)
LOW_CONFIDENCE_FONT = Font(name="Arial", size=11, color="CC6600")

EMAIL_MISMATCH_FILL = PatternFill(
    start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
)
EMAIL_MISMATCH_FONT = Font(name="Arial", size=11, color="B71C1C")

ROLE_UNMAPPED_FILL = PatternFill(
    start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"
)

NORMAL_FONT = Font(name="Arial", size=11)
NORMAL_ALIGNMENT = Alignment(
    horizontal="left", vertical="center", wrap_text=True
)


SSO_HEADERS = [
    "STT", "Ho va ten", "Ma chi nhanh", "Ten Chi nhanh", "User IPCAS", "So CCCD",
    "Email tai Agribank", "SDT", "Phan quyen", "Ma lien ngan hang",
    "Vai tro (goi y)",
]

# OCR grid col index -> Excel export column (mẫu 10 cột mới).
_SSO_GRID_TO_EXCEL_COL_10 = {i: i for i in range(10)}

# Mẫu cũ 9 cột -> cột Excel mới (tách Phòng/Đơn vị thành Mã CN + Tên CN khi export).
_SSO_GRID_TO_EXCEL_COL_9 = {
    0: 0, 1: 1, 2: 2, 3: 4, 4: 5, 5: 6, 6: 7, 7: 8, 8: 9,
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)




def _export_sso_sheet(ws, table, threshold: float, start_row: int = 1) -> int:
    from app.services.email_reconcile import email_mismatch_with_ipcas, email_needs_review
    from app.services.user_mapping import (
        _parse_branch_code_digits,
        _parse_department_cell,
        normalize_roles,
    )

    grid = {}
    for cell in table.cells:
        grid[(cell.row, cell.col)] = (cell.text, cell.confidence)
    is_new_layout = table.num_cols >= 10
    grid_to_excel = _SSO_GRID_TO_EXCEL_COL_10 if is_new_layout else _SSO_GRID_TO_EXCEL_COL_9
    email_grid_col = 6 if is_new_layout else 5
    role_grid_col = 8 if is_new_layout else 7
    ipcas_grid_col = 4 if is_new_layout else 3

    row = start_row
    for c, title in enumerate(SSO_HEADERS):
        cell = ws.cell(row=row, column=c + 1)
        cell.value = title
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1
    data_rows = max(table.num_rows, 1)
    for r in range(data_rows):
        dept_text, _ = grid.get((r, 2), ("", 1.0))
        branch_code = (
            grid.get((r, 2), ("", 1.0))[0].strip()
            if is_new_layout
            else _parse_branch_code_digits(dept_text)
        )
        ipcas = grid.get((r, ipcas_grid_col), ("", 1.0))[0]
        email_text, email_conf = grid.get((r, email_grid_col), ("", 1.0))
        role_text, role_conf = grid.get((r, role_grid_col), ("", 1.0))
        role_suggested = ";".join(normalize_roles(role_text)) if role_text else ""
        email_mismatch = bool(ipcas and email_text and email_mismatch_with_ipcas(email_text, ipcas))
        email_uncertain = bool(email_text and email_needs_review(email_text))
        role_unmapped = bool(role_text.strip() and not role_suggested)

        for excel_c in range(len(SSO_HEADERS)):
            ws_cell = ws.cell(row=row, column=excel_c + 1)
            ws_cell.alignment = NORMAL_ALIGNMENT
            ws_cell.border = THIN_BORDER

            if excel_c == 2 and not is_new_layout:
                _, bc, _ = _parse_department_cell(dept_text)
                ws_cell.value = bc or branch_code
                ws_cell.font = NORMAL_FONT
                continue
            if excel_c == 3 and not is_new_layout:
                _, _, bn = _parse_department_cell(dept_text)
                ws_cell.value = bn or dept_text
                ws_cell.font = NORMAL_FONT
                continue
            if excel_c == 10:
                ws_cell.value = role_suggested
                if role_unmapped:
                    ws_cell.fill = ROLE_UNMAPPED_FILL
                    ws_cell.font = LOW_CONFIDENCE_FONT
                else:
                    ws_cell.font = NORMAL_FONT
                continue

            grid_c = next((g for g, e in grid_to_excel.items() if e == excel_c), None)
            if grid_c is None:
                continue
            text_val, conf = grid.get((r, grid_c), ("", 1.0))
            ws_cell.value = text_val
            if grid_c == email_grid_col and (email_mismatch or email_uncertain):
                ws_cell.fill = EMAIL_MISMATCH_FILL
                ws_cell.font = EMAIL_MISMATCH_FONT
            elif grid_c == role_grid_col and role_unmapped:
                ws_cell.fill = ROLE_UNMAPPED_FILL
                ws_cell.font = LOW_CONFIDENCE_FONT
            elif conf < threshold:
                ws_cell.fill = LOW_CONFIDENCE_FILL
                ws_cell.font = LOW_CONFIDENCE_FONT
            else:
                ws_cell.font = NORMAL_FONT
        row += 1
    return row


def _norm_excel_header(val: str) -> str:
    return _normalize_header_key(val)


def _is_sso_header_row(row: list[str]) -> bool:
    aliases = {
        "stt", "ho va ten", "ma chi nhanh", "ten chi nhanh", "user ipcas", "so cccd",
        "email", "email tai agribank", "sdt", "so dien thoai", "phan quyen",
        "ma lien ngan hang", "ghi chu / ma dv",
        # Mẫu cũ
        "phong/don vi", "ma cn",
    }
    hits = sum(1 for cell in row[:12] if _norm_excel_header(cell) in aliases)
    return hits >= 4


def export_to_excel(
    result: OcrResult,
    *,
    page_numbers: list[int] | None = None,
) -> Path:
    """
    Export OCR result to an Excel file.

    Creates one sheet per page. Each table on a page is placed
    sequentially with a gap row between tables.

    Args:
        result: The OcrResult to export
        page_numbers: Optional subset of page numbers to include (1-based)

    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    threshold = settings.ocr_confidence_threshold
    page_filter = set(page_numbers) if page_numbers else None
    pages = result.pages
    if page_filter:
        pages = [p for p in result.pages if p.page_number in page_filter]
    if not pages:
        raise ValueError("Không có trang nào để xuất Excel")

    for page in pages:
        sheet_name = f"Trang {page.page_number}"
        ws = wb.create_sheet(title=sheet_name)

        current_row = 1

        sso_tables = [t for t in page.tables if t.table_kind == "sso_agribank"]
        if sso_tables:
            current_row = 1
            for table in sso_tables:
                if not table.cells:
                    continue
                current_row = _export_sso_sheet(ws, table, threshold, start_row=current_row) + 2
            _auto_adjust_column_widths(ws)
            continue

        for table in page.tables:
            if not table.cells:
                continue

            # Write table title
            ws.cell(row=current_row, column=1).value = (
                f"Bảng {table.table_index + 1} — "
                f"{table.num_rows} dòng × {table.num_cols} cột"
            )
            ws.cell(row=current_row, column=1).font = Font(
                name="Arial", size=12, bold=True, color="1F4E79"
            )
            current_row += 1

            # Build a 2D grid from cell data
            grid: dict[tuple[int, int], tuple[str, float]] = {}
            for cell in table.cells:
                grid[(cell.row, cell.col)] = (cell.text, cell.confidence)

            # Write cells
            for row_idx in range(table.num_rows):
                for col_idx in range(table.num_cols):
                    text, confidence = grid.get((row_idx, col_idx), ("", 1.0))

                    excel_row = current_row + row_idx
                    excel_col = col_idx + 1

                    ws_cell = ws.cell(row=excel_row, column=excel_col)
                    ws_cell.value = text
                    ws_cell.alignment = NORMAL_ALIGNMENT
                    ws_cell.border = THIN_BORDER

                    # First row as header
                    if row_idx == 0:
                        ws_cell.fill = HEADER_FILL
                        ws_cell.font = HEADER_FONT
                    elif confidence < threshold:
                        # Highlight low-confidence cells
                        ws_cell.fill = LOW_CONFIDENCE_FILL
                        ws_cell.font = LOW_CONFIDENCE_FONT
                    else:
                        ws_cell.font = NORMAL_FONT

            current_row += table.num_rows + 2  # Gap between tables

        # Auto-adjust column widths
        _auto_adjust_column_widths(ws)

    # Add a legend sheet if there are low-confidence cells
    _add_legend_sheet(wb, threshold)

    # Save
    base = result.filename.rsplit(".", 1)[0]
    if page_filter:
        page_tag = "-".join(str(p) for p in sorted(page_filter))
        filename = f"{result.job_id}_{base}_trang-{page_tag}.xlsx"
    else:
        filename = f"{result.job_id}_{base}.xlsx"
    export_file = settings.export_path / filename

    wb.save(str(export_file))
    logger.info("Excel exported: %s", export_file)

    return export_file


def import_from_excel(
    excel_path: str | Path,
    job_id: str,
    filename: str,
) -> OcrResult:
    """
    Import data from Excel into OcrResult structure.

    Supported formats:
    - Any sheet with a plain rectangular table (non-empty used range)
    - Exported OCR workbook format where each table starts with:
      "Bảng X — ...", followed by the table grid, with blank rows between tables

    File có công thức: cố gắng materialize giá trị (Excel COM / formulas /
    dual-read); ô công thức chưa tính được sẽ để trống + warnings.
    """
    excel_path = Path(excel_path)
    import_warnings: list[str] = []
    materialize_path, mat_warnings = _materialize_formula_values(excel_path)
    import_warnings.extend(mat_warnings)

    try:
        wb_data = load_workbook(filename=str(materialize_path), data_only=True)
        wb_form = load_workbook(filename=str(excel_path), data_only=False)
    finally:
        if materialize_path != excel_path and materialize_path.exists():
            try:
                materialize_path.unlink()
            except OSError:
                pass

    pages: list[PageResult] = []
    formula_empty_total = 0
    page_number = 1
    for ws_data in wb_data.worksheets:
        title = ws_data.title.strip().lower()
        if title == "chú thích":
            continue
        ws_form = None
        try:
            ws_form = wb_form[ws_data.title]
        except KeyError:
            ws_form = None

        tables, empty_count = _extract_tables_from_sheet(ws_data, ws_form)
        formula_empty_total += empty_count
        if not tables:
            continue

        pages.append(
            PageResult(
                page_number=page_number,
                image_path="",
                tables=tables,
                raw_text="",
            )
        )
        page_number += 1

    if not pages:
        raise ValueError("Không tìm thấy dữ liệu bảng trong file Excel")

    if formula_empty_total:
        import_warnings.append(
            f"{formula_empty_total} ô công thức chưa lấy được giá trị "
            "(mở file bằng Excel → Save, hoặc Paste Values rồi upload lại)."
        )

    now = datetime.now()
    return OcrResult(
        job_id=job_id,
        filename=filename,
        total_pages=len(pages),
        pages=pages,
        is_complete=True,
        warnings=import_warnings,
        created_at=now,
        updated_at=now,
    )


def _materialize_formula_values(excel_path: Path) -> tuple[Path, list[str]]:
    """
    Tạo bản workbook đã tính công thức (nếu được).

    Thứ tự: Excel COM (Windows) → thư viện formulas (nếu cài) → giữ file gốc.
    """
    warnings: list[str] = []
    if not _workbook_has_formulas(excel_path):
        return excel_path, warnings

    com_path = _materialize_via_excel_com(excel_path)
    if com_path is not None:
        warnings.append("Đã tính công thức Excel qua Microsoft Excel (COM).")
        return com_path, warnings

    formulas_path = _materialize_via_formulas_lib(excel_path)
    if formulas_path is not None:
        warnings.append("Đã tính công thức Excel qua thư viện formulas.")
        return formulas_path, warnings

    warnings.append(
        "File có công thức nhưng chưa cache giá trị. "
        "Hệ thống đọc data_only; ô chưa tính có thể trống. "
        "Khuyến nghị: mở Excel → Ctrl+Shift+End → Copy → Paste Values → Save."
    )
    return excel_path, warnings


def _workbook_has_formulas(excel_path: Path) -> bool:
    try:
        wb = load_workbook(filename=str(excel_path), data_only=False, read_only=True)
    except Exception:
        return False
    try:
        for ws in wb.worksheets:
            if ws.title.strip().lower() == "chú thích":
                continue
            for row in ws.iter_rows(max_row=min(ws.max_row or 1, 500), max_col=min(ws.max_column or 1, 30)):
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        return True
        return False
    finally:
        wb.close()


def _materialize_via_excel_com(excel_path: Path) -> Path | None:
    """Windows + Excel + pywin32: CalculateFull rồi SaveCopyAs."""
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError:
        return None

    out = excel_path.with_name(f"{excel_path.stem}__values{excel_path.suffix}")
    excel = None
    wb = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(excel_path.resolve()))
        excel.CalculateFullRebuild()
        wb.SaveCopyAs(str(out.resolve()))
        wb.Close(SaveChanges=False)
        wb = None
        return out if out.exists() else None
    except Exception as exc:
        logger.info("Excel COM materialize skipped: %s", exc)
        if out.exists():
            try:
                out.unlink()
            except OSError:
                pass
        return None
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _materialize_via_formulas_lib(excel_path: Path) -> Path | None:
    """Optional dependency: formulas (pip install formulas)."""
    try:
        import formulas  # type: ignore
    except ImportError:
        return None

    out_dir = excel_path.parent / f".excel_calc_{excel_path.stem}"
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        xl_model = formulas.ExcelModel().loads(str(excel_path)).finish()
        xl_model.calculate()
        xl_model.write(dirpath=str(out_dir))
        candidates = list(out_dir.glob("*.xlsx")) + list(out_dir.glob("*.xlsm"))
        if not candidates:
            return None
        preferred = out_dir / excel_path.name
        chosen = preferred if preferred.exists() else candidates[0]
        out = excel_path.with_name(f"{excel_path.stem}__values{excel_path.suffix}")
        if out.exists():
            out.unlink()
        chosen.replace(out)
        return out if out.exists() else None
    except Exception as exc:
        logger.info("formulas lib materialize skipped: %s", exc)
        return None
    finally:
        try:
            if out_dir.exists():
                for p in out_dir.rglob("*"):
                    if p.is_file():
                        p.unlink()
                try:
                    out_dir.rmdir()
                except OSError:
                    pass
        except OSError:
            pass


def _extract_tables_from_sheet(
    ws_data, ws_form=None
) -> tuple[list[TableData], int]:
    """Extract one or many tables from a worksheet. Returns (tables, formula_empty_count)."""
    title_rows = _find_table_title_rows(ws_form or ws_data)
    tables: list[TableData] = []
    formula_empty = 0

    if title_rows:
        for idx, title_row in enumerate(title_rows):
            start_row = title_row + 1
            end_row = (
                title_rows[idx + 1] - 2
                if idx + 1 < len(title_rows)
                else (ws_data.max_row or 1)
            )
            matrix, empty = _read_matrix(ws_data, start_row, end_row, ws_form=ws_form)
            formula_empty += empty
            if not matrix:
                continue
            tables.append(_matrix_to_table(matrix, len(tables)))
        return tables, formula_empty

    # Fallback: parse whole used range as one table
    matrix, empty = _read_matrix(ws_data, 1, ws_data.max_row or 1, ws_form=ws_form)
    formula_empty += empty
    if matrix:
        tables.append(_matrix_to_table(matrix, 0))
    return tables, formula_empty


def _find_table_title_rows(ws) -> list[int]:
    """Find rows that look like exported table titles: 'Bảng X — ...'."""
    rows: list[int] = []
    for r in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and re.match(r"^\s*Bảng\s+\d+", v.strip(), re.IGNORECASE):
            rows.append(r)
    return rows


def _excel_cell_text(v) -> str:
    """Normalize Excel cell value to string (whole-number floats -> int text)."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        if v.is_integer() or abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return str(v).strip()
    if isinstance(v, int):
        return str(v)
    text = str(v).strip()
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?[eE][+-]?\d+", text):
        try:
            num = float(text)
            if abs(num - round(num)) < 1e-6:
                return str(int(round(num)))
        except ValueError:
            pass
    return text


def _build_merge_anchor_map(ws) -> dict[tuple[int, int], tuple[int, int]]:
    """Map (row,col) trong vùng merge → (top,left) anchor."""
    mapping: dict[tuple[int, int], tuple[int, int]] = {}
    try:
        ranges = list(ws.merged_cells.ranges)
    except Exception:
        return mapping
    for mr in ranges:
        min_r, min_c, max_r, max_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    mapping[(r, c)] = (min_r, min_c)
    return mapping


def _read_matrix(
    ws_data,
    start_row: int,
    end_row: int,
    *,
    ws_form=None,
) -> tuple[list[list[str]], int]:
    """Read non-empty rectangular matrix from row range.

    Ưu tiên giá trị data_only; fallback formula sheet / ô merge.
    Trả (matrix, số ô công thức trống).
    """
    rows: list[list[str]] = []
    max_col = 0
    formula_empty = 0
    form = ws_form or ws_data
    merge_map = _build_merge_anchor_map(form)
    max_column = max(ws_data.max_column or 1, getattr(form, "max_column", 1) or 1)

    for r in range(start_row, end_row + 1):
        vals: list[str] = []
        row_non_empty = False
        for c in range(1, max_column + 1):
            raw = ws_data.cell(row=r, column=c).value
            form_raw = form.cell(row=r, column=c).value if form is not ws_data else raw
            is_formula = isinstance(form_raw, str) and form_raw.startswith("=")

            if raw is None and is_formula:
                # Chưa có cache — thử resolve đơn giản từ ô tham chiếu gần
                resolved = _try_resolve_simple_formula(form_raw, ws_data, form, r, c)
                if resolved is not None:
                    raw = resolved
                else:
                    formula_empty += 1
            elif raw is None and form_raw is not None and not is_formula:
                raw = form_raw

            if (raw is None or raw == "") and (r, c) in merge_map:
                ar, ac = merge_map[(r, c)]
                raw = ws_data.cell(row=ar, column=ac).value
                if raw is None:
                    anchor_form = form.cell(row=ar, column=ac).value
                    if not (isinstance(anchor_form, str) and anchor_form.startswith("=")):
                        raw = anchor_form

            txt = _excel_cell_text(raw)
            vals.append(txt)
            if txt:
                row_non_empty = True
                max_col = max(max_col, c)
        if row_non_empty:
            rows.append(vals)

    if not rows or max_col == 0:
        return [], formula_empty

    # Giữ tối thiểu số cột header SSO (tránh cắt cột cuối → nhầm layout 9 cột).
    header_cols = 0
    for row in rows[:3]:
        filled = sum(1 for x in row if str(x or "").strip())
        if filled >= 6:
            header_cols = max(header_cols, filled)
    keep_cols = max(max_col, header_cols)
    if header_cols >= 10:
        keep_cols = max(keep_cols, 10)

    matrix = []
    for r in rows:
        trimmed = r[:keep_cols]
        if len(trimmed) < keep_cols:
            trimmed = trimmed + [""] * (keep_cols - len(trimmed))
        matrix.append(trimmed)
    return matrix, formula_empty


_CELL_REF_RE = re.compile(r"^\s*=?\s*([A-Za-z]{1,3})(\d+)\s*$")
_CONCAT_EDGE_RE = re.compile(
    r'^\s*=\s*(?:LOWER|UPPER|TRIM)\s*\(\s*([A-Za-z]{1,3})(\d+)\s*\)\s*&\s*"([^"]*)"\s*$',
    re.IGNORECASE,
)
_AMP_LITERAL_RE = re.compile(
    r'^\s*=\s*([A-Za-z]{1,3})(\d+)\s*&\s*"([^"]*)"\s*$',
    re.IGNORECASE,
)
_IF_EMPTY_RE = re.compile(
    r'^\s*=\s*IF\s*\(\s*([A-Za-z]{1,3})(\d+)\s*=\s*""\s*,\s*""\s*,\s*(.+)\s*\)\s*$',
    re.IGNORECASE,
)


def _col_letters_to_index(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _try_resolve_simple_formula(formula: str, ws_data, ws_form, row: int, col: int):
    """Resolve một số công thức SSO phổ biến khi thiếu cache."""
    f = (formula or "").strip()
    if not f.startswith("="):
        return None

    m = _IF_EMPTY_RE.match(f)
    if m:
        ref_c, ref_r = _col_letters_to_index(m.group(1)), int(m.group(2))
        seed = ws_data.cell(row=ref_r, column=ref_c).value
        if seed is None:
            seed_f = ws_form.cell(row=ref_r, column=ref_c).value
            if not (isinstance(seed_f, str) and seed_f.startswith("=")):
                seed = seed_f
        if seed is None or _excel_cell_text(seed) == "":
            return ""
        inner = m.group(3).strip()
        if not inner.startswith("="):
            inner = "=" + inner
        return _try_resolve_simple_formula(inner, ws_data, ws_form, row, col)

    m = _CONCAT_EDGE_RE.match(f)
    if m:
        ref_c, ref_r = _col_letters_to_index(m.group(1)), int(m.group(2))
        seed = ws_data.cell(row=ref_r, column=ref_c).value
        if seed is None:
            seed_f = ws_form.cell(row=ref_r, column=ref_c).value
            if not (isinstance(seed_f, str) and seed_f.startswith("=")):
                seed = seed_f
        text = _excel_cell_text(seed)
        if not text:
            return None
        fn = f.upper()
        if "LOWER" in fn:
            text = text.lower()
        elif "UPPER" in fn:
            text = text.upper()
        return text + m.group(3)

    m = _AMP_LITERAL_RE.match(f)
    if m:
        ref_c, ref_r = _col_letters_to_index(m.group(1)), int(m.group(2))
        seed = ws_data.cell(row=ref_r, column=ref_c).value
        if seed is None:
            seed_f = ws_form.cell(row=ref_r, column=ref_c).value
            if not (isinstance(seed_f, str) and seed_f.startswith("=")):
                seed = seed_f
        text = _excel_cell_text(seed)
        if not text:
            return None
        return text + m.group(3)

    m = _CELL_REF_RE.match(f)
    if m:
        ref_c, ref_r = _col_letters_to_index(m.group(1)), int(m.group(2))
        seed = ws_data.cell(row=ref_r, column=ref_c).value
        if seed is None:
            seed_f = ws_form.cell(row=ref_r, column=ref_c).value
            if isinstance(seed_f, str) and seed_f.startswith("="):
                return _try_resolve_simple_formula(seed_f, ws_data, ws_form, ref_r, ref_c)
            seed = seed_f
        return seed

    m = re.match(
        r'^\s*=\s*(LOWER|UPPER|TRIM)\s*\(\s*([A-Za-z]{1,3})(\d+)\s*\)\s*$',
        f,
        re.IGNORECASE,
    )
    if m:
        ref_c, ref_r = _col_letters_to_index(m.group(2)), int(m.group(3))
        seed = ws_data.cell(row=ref_r, column=ref_c).value
        if seed is None:
            seed_f = ws_form.cell(row=ref_r, column=ref_c).value
            if not (isinstance(seed_f, str) and seed_f.startswith("=")):
                seed = seed_f
        text = _excel_cell_text(seed)
        fn = m.group(1).upper()
        if fn == "LOWER":
            return text.lower()
        if fn == "UPPER":
            return text.upper()
        return text.strip()

    return None


def _matrix_to_table(matrix: list[list[str]], table_index: int) -> TableData:
    matrix = _strip_sso_preamble_rows(matrix)
    num_rows = len(matrix)
    num_cols = max((len(r) for r in matrix), default=0)
    cells: list[CellData] = []

    for r, row_vals in enumerate(matrix):
        padded = row_vals + [""] * (num_cols - len(row_vals))
        for c, txt in enumerate(padded):
            cells.append(
                CellData(
                    row=r,
                    col=c,
                    text=txt,
                    confidence=1.0,
                    bbox=[],
                )
            )

    table_kind = "sso_agribank" if num_cols >= 6 else ""
    return TableData(
        table_index=table_index,
        num_rows=num_rows,
        num_cols=num_cols,
        cells=cells,
        html="",
        table_kind=table_kind,
    )


def _auto_adjust_column_widths(ws) -> None:
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # Vietnamese text can be wider
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
        # Add some padding, cap at 50
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 12)


def _add_legend_sheet(wb: Workbook, threshold: float) -> None:
    """Add a legend/info sheet explaining the highlighting."""
    ws = wb.create_sheet(title="Chú thích", index=0)

    ws.cell(row=1, column=1).value = "Chú thích — Kết quả OCR"
    ws.cell(row=1, column=1).font = Font(
        name="Arial", size=14, bold=True, color="1F4E79"
    )

    ws.cell(row=3, column=1).value = "Màu ô"
    ws.cell(row=3, column=2).value = "Ý nghĩa"
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=3, column=2).font = Font(bold=True)

    # Header row example
    ws.cell(row=4, column=1).fill = HEADER_FILL
    ws.cell(row=4, column=1).font = HEADER_FONT
    ws.cell(row=4, column=1).value = "Header"
    ws.cell(row=4, column=2).value = "Dòng tiêu đề bảng"

    # Low confidence example
    ws.cell(row=5, column=1).fill = LOW_CONFIDENCE_FILL
    ws.cell(row=5, column=1).font = LOW_CONFIDENCE_FONT
    ws.cell(row=5, column=1).value = "Cần review"
    ws.cell(row=5, column=2).value = (
        f"Ô có độ tin cậy OCR < {threshold:.0%} — cần kiểm tra lại"
    )

    # Normal example
    ws.cell(row=6, column=1).font = NORMAL_FONT
    ws.cell(row=6, column=1).value = "Bình thường"
    ws.cell(row=6, column=2).value = "Ô đã nhận dạng chính xác"

    ws.cell(row=8, column=1).value = (
        "Lưu ý: Các ô được highlight vàng cần được kiểm tra lại "
        "trước khi sử dụng để tạo lô user."
    )
    ws.cell(row=8, column=1).font = Font(italic=True, color="666666")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
