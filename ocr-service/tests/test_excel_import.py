"""Tests for Excel import — formulas, floats, skip warnings."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.services.excel_service import (
    _excel_cell_text,
    _try_resolve_simple_formula,
    import_from_excel,
)
from app.services.user_mapping import map_result_to_users, map_table_to_users


def _write_sso_book(path: Path, *, with_formulas: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DS"
    headers = [
        "STT",
        "Ho va ten",
        "Ma chi nhanh",
        "Ten Chi nhanh",
        "User IPCAS",
        "So CCCD",
        "Email tai Agribank",
        "SDT",
        "Phan quyen",
        "Ma lien ngan hang",
    ]
    ws.append(headers)
    # Row 2 — plain values
    ws.append([
        1,
        "Nguyen Van A",
        1500,
        "CN Test",
        "HQPTESTA",
        123456789012,
        "hqptesta@agribank.com.vn",
        "0912345678",
        "Dai ly vien",
        "95204001",
    ])
    # Row 3 — email/phone via formula referencing IPCAS / literal
    ws.append([
        2,
        "Tran Thi B",
        1500,
        "CN Test",
        "HQPTESTB",
        987654321098.0,  # float CCCD
        '=LOWER(E3)&"@agribank.com.vn"' if with_formulas else "hqptestb@agribank.com.vn",
        "0987654321",
        "Ke toan vien",
        "95204001",
    ])
    # Row 4 — missing identity (no email, no ipcas) should be skipped when mapped
    ws.append([
        3,
        "Le Van C",
        1500,
        "CN Test",
        "",
        "",
        '=IF(E4="","",LOWER(E4)&"@agribank.com.vn")' if with_formulas else "",
        "",
        "",
        "",
    ])
    wb.save(path)


def test_excel_cell_text_float_and_sci():
    assert _excel_cell_text(123456789012.0) == "123456789012"
    assert _excel_cell_text("1.23456789012E+11") == "123456789012"
    assert _excel_cell_text(None) == ""


def test_try_resolve_simple_lower_concat(tmp_path: Path):
    path = tmp_path / "f.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["E2"] = "HQPTEST"
    ws["G2"] = '=LOWER(E2)&"@agribank.com.vn"'
    wb.save(path)

    from openpyxl import load_workbook

    wb_data = load_workbook(path, data_only=True)
    wb_form = load_workbook(path, data_only=False)
    ws_d, ws_f = wb_data.active, wb_form.active
    # data_only without Excel cache → None
    assert ws_d["G2"].value is None
    resolved = _try_resolve_simple_formula(
        ws_f["G2"].value, ws_d, ws_f, 2, 7
    )
    assert resolved == "hqptest@agribank.com.vn"


def test_import_excel_resolves_simple_email_formula(tmp_path: Path):
    path = tmp_path / "sso_formula.xlsx"
    _write_sso_book(path, with_formulas=True)
    result = import_from_excel(path, "job1", "sso_formula.xlsx")
    assert result.pages
    users, warnings = map_result_to_users(result)
    emails = {u.email for u in users}
    assert "hqptesta@agribank.com.vn" in emails
    assert "hqptestb@agribank.com.vn" in emails
    # row without identity skipped
    assert all(u.email for u in users)
    assert any("bỏ qua" in w.lower() or "bo qua" in w.lower() or "thiếu" in w.lower() or "thieu" in w.lower() for w in warnings)


def test_import_excel_plain_values_and_cccd_float(tmp_path: Path):
    path = tmp_path / "sso_plain.xlsx"
    _write_sso_book(path, with_formulas=False)
    result = import_from_excel(path, "job2", "sso_plain.xlsx")
    users, _warnings = map_result_to_users(result)
    by_email = {u.email: u for u in users}
    assert by_email["hqptestb@agribank.com.vn"].cccd == "987654321098"


def test_truncated_9col_still_maps_branch_code():
    """Paste Values cắt cột cuối → 9 cột nhưng vẫn map đúng mã CN (layout 10)."""
    from app.models.schemas import CellData, TableData
    from app.services.user_mapping import map_table_to_users

    # Không có header (đã strip): STT, tên, mã CN, tên CN, IPCAS, CCCD, email, phone, role
    rows = [
        ["1", "Nguyen Van A", "1500", "CN Ha Noi", "HQPAAA", "001234567890", "hqpaaa@agribank.com.vn", "0911111111", "Dai ly vien"],
        ["2", "Tran Thi B", "1600", "CN Hue", "HQPBBB", "001234567891", "hqpbbb@agribank.com.vn", "0922222222", "Ke toan vien"],
        ["3", "Le Van C", "1700", "CN Da Nang", "HQPCCC", "001234567892", "hqpccc@agribank.com.vn", "0933333333", "Dai ly vien"],
    ]
    cells = []
    for r, row in enumerate(rows):
        for c, txt in enumerate(row):
            cells.append(CellData(row=r, col=c, text=txt, confidence=1.0, bbox=[]))
    table = TableData(
        table_index=0,
        num_rows=len(rows),
        num_cols=9,
        cells=cells,
        table_kind="sso_agribank",
    )
    users, _w = map_table_to_users(table)
    assert len(users) == 3
    assert users[0].branch_code == "1500"
    assert users[1].branch_code == "1600"
    assert users[0].ipcas_code == "HQPAAA"


def test_header_ma_chi_nhanh_maps_branch_code():
    from app.services.user_mapping import _map_header

    header = [
        "STT",
        "Ho va ten",
        "Ma chi nhanh",
        "Ten Chi nhanh",
        "User IPCAS",
        "So CCCD",
        "Email tai Agribank",
        "SDT",
        "Phan quyen",
        "Ma lien ngan hang",
    ]
    col_map = _map_header(header)
    assert col_map.get("branch_code") == 2
    assert col_map.get("branch_name") == 3
